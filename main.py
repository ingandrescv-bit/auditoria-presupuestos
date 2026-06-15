from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import io, zipfile, tempfile, os, math

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

MAX_ABS = 50_000_000_000

# ── helpers ──────────────────────────────────────────────────
def to_num(v):
    if v is None or v == "": return None
    try: return float(v)
    except: return None

def chk(a, b, res, tol=0.02):
    if not a or not b or res is None: return None
    esp = a * b
    diff = abs(esp - res)
    return (esp, diff) if diff > max(abs(esp)*tol, 500) else None

def fill(hex6): return PatternFill("solid", fgColor="FF" + hex6.upper())
def fnt(hex6, bold=False, sz=10): return Font(color="FF"+hex6.upper(), bold=bold, size=sz, name="Calibri")
def aln(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Lee estilos de cada celda desde openpyxl ────────────────
def leer_estilos_xlsx(ws_orig):
    """Devuelve dict {(row,col): {bg, fg, bold, italic, sz, h_align, v_align, wrap, fmt}}"""
    estilos = {}
    for row in ws_orig.iter_rows():
        for cell in row:
            r, c = cell.row, cell.column
            try:
                bg = None
                if cell.fill and cell.fill.fill_type == "solid":
                    argb = str(cell.fill.fgColor.rgb)
                    if argb and argb != "00000000" and argb.lower() != "ffffffff":
                        bg = argb[-6:]  # strip alpha

                fg = "000000"
                if cell.font and cell.font.color:
                    try:
                        argb = str(cell.font.color.rgb)
                        if argb and len(argb) >= 6:
                            fg = argb[-6:]
                    except: pass

                estilos[(r,c)] = {
                    "bg": bg,
                    "fg": fg,
                    "bold": cell.font.bold if cell.font else False,
                    "italic": cell.font.italic if cell.font else False,
                    "sz": cell.font.size if cell.font and cell.font.size else 10,
                    "h": cell.alignment.horizontal if cell.alignment else "left",
                    "v": cell.alignment.vertical if cell.alignment else "center",
                    "wrap": cell.alignment.wrap_text if cell.alignment else False,
                    "fmt": cell.number_format if cell.number_format else "General",
                }
            except:
                estilos[(r,c)] = {"bg": None, "fg":"000000","bold":False,"italic":False,"sz":10,"h":"left","v":"center","wrap":False,"fmt":"General"}
    return estilos

def leer_estilos_numbers(path, sheet_name):
    """Lee estilos desde archivo .numbers usando numbers_parser"""
    from numbers_parser import Document
    doc = Document(path)
    estilos = {}
    for sheet in doc.sheets:
        if sheet.name != sheet_name: continue
        table = sheet.tables[0]
        for ri, row in enumerate(table.iter_rows()):
            for ci, cell in enumerate(row):
                r, c = ri+1, ci+1
                try:
                    sty = cell.style
                    bg = None
                    if sty and sty.bg_color:
                        bc = sty.bg_color
                        hex6 = f"{bc.r:02X}{bc.g:02X}{bc.b:02X}"
                        if hex6 not in ("FFFFFF","000000"):
                            bg = hex6
                    fg = "000000"
                    if sty and sty.font_color:
                        fc = sty.font_color
                        fg = f"{fc.r:02X}{fc.g:02X}{fc.b:02X}"
                    h_map = {0:"left",1:"right",2:"center"}
                    v_map = {0:"top",1:"center"}
                    h_align = h_map.get(sty.alignment.horizontal.value if sty else 0, "left")
                    v_align = v_map.get(sty.alignment.vertical.value if sty else 1, "center")
                    estilos[(r,c)] = {
                        "bg": bg,
                        "fg": fg,
                        "bold": sty.bold if sty else False,
                        "italic": sty.italic if sty else False,
                        "sz": sty.font_size if sty else 10,
                        "h": h_align,
                        "v": v_align,
                        "wrap": sty.text_wrap if sty else False,
                        "fmt": "General",
                    }
                except:
                    estilos[(r,c)] = {"bg":None,"fg":"000000","bold":False,"italic":False,"sz":10,"h":"left","v":"center","wrap":False,"fmt":"General"}
        break
    return estilos

# ── Detecta errores en los datos ────────────────────────────
def detectar_errores(all_rows, header_col_count=8):
    errMap = {}
    def add(i, col, tipo, msg, correcto=None):
        if i not in errMap: errMap[i] = []
        errMap[i].append({"col":col,"tipo":tipo,"msg":msg,"correcto":correcto})

    for i, f in enumerate(all_rows):
        if i == 0 or not f: continue
        desc = str(f[1] if len(f)>1 else "").strip()
        if not desc: continue

        cant = to_num(f[3] if len(f)>3 else None)
        vu   = to_num(f[4] if len(f)>4 else None)
        fac  = to_num(f[5] if len(f)>5 else None)
        vc   = to_num(f[6] if len(f)>6 else None)
        sub  = to_num(f[7] if len(f)>7 else None)

        # Valores en columnas extra (mas alla de las 8 esperadas)
        for ci in range(header_col_count, len(f)):
            v = to_num(f[ci])
            if v is not None and abs(v) > 1000:
                add(i, ci, "extra", f"Valor inesperado en col {ci+1}: {v:,.0f}")

        # Subtotal absurdo
        if sub is not None and abs(sub) > MAX_ABS:
            corr = cant*vc if cant and vc else None
            add(i, 7, "absurdo", f"SUBTOTAL ABSURDO: {sub:,.0f}", corr)
            continue

        if not vu or vu < 100: continue

        # VU x factor != V.con AIU
        if vu and fac and vc:
            r = chk(vu, fac, vc)
            if r:
                esp, diff = r
                add(i, 6, "formula", f"V.AIU: {vu:.0f}x{fac}={esp:,.0f} != {vc:,.0f}", esp)

        # cant x vcAIU != subtotal
        if cant and vc and sub:
            r2 = chk(cant, vc, sub)
            if r2:
                esp2, diff2 = r2
                add(i, 7, "formula", f"Subtotal: {cant}x{vc:,.0f}={esp2:,.0f} != {sub:,.0f}", esp2)

    return errMap

# ── Calcula resumen ejecutivo ────────────────────────────────
def calcular_resumen(all_rows, errMap):
    import datetime
    total = 0; caps = {}; cap_act = None
    for i, f in enumerate(all_rows):
        if i == 0 or not f: continue
        vu = to_num(f[4] if len(f)>4 else None)
        sub = to_num(f[7] if len(f)>7 else None)
        desc = str(f[1] if len(f)>1 else "").strip()
        dl = desc.lower()
        es_cap = (vu is None or vu < 50)
        if es_cap and desc and "subtotal" not in dl and "total" not in dl:
            cap_act = desc
        if not es_cap and vu and vu >= 100 and sub and abs(sub) < MAX_ABS:
            if i not in errMap or not any(e["tipo"]=="absurdo" for e in errMap[i]):
                total += sub
                if cap_act: caps[cap_act] = caps.get(cap_act, 0) + sub

    criticos = sum(1 for v in errMap.values() for e in v if e["tipo"]=="absurdo")
    formulas = sum(1 for v in errMap.values() for e in v if e["tipo"]=="formula")
    extras   = sum(1 for v in errMap.values() for e in v if e["tipo"]=="extra")
    n_err    = criticos + formulas + extras
    puntaje  = max(0, 100 - criticos*30 - formulas*15 - extras*5)
    area     = 858.35
    fecha    = datetime.date.today().strftime("%d/%m/%Y")

    return {"total":total,"caps":caps,"criticos":criticos,"formulas":formulas,
            "extras":extras,"n_err":n_err,"puntaje":puntaje,"area":area,"fecha":fecha}

# ── Escribe Excel con estilos originales + auditoria ────────
def escribir_excel(all_rows, estilos_orig, errMap, resumen, sheet_name):
    OFFSET = 5  # filas de resumen ejecutivo antes del presupuesto

    # Colores semaforo (no cambian)
    F_ROJO    = fill("C00000"); FNT_ROJO    = fnt("FFFFFF", True)
    F_NARANJA = fill("E26B0A"); FNT_NARANJA = fnt("FFFFFF", True)
    F_AMARILLO= fill("FFD700"); FNT_AMARILLO= fnt("252525", True)
    F_VERDE_OK= fill("375623"); FNT_VERDE_OK= fnt("FFFFFF", True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Hoja1")[:31]

    tot = resumen["total"]; area = resumen["area"]
    fecha = resumen["fecha"]

    # ── RESUMEN EJECUTIVO ──────────────────────────────────────
    def sc(r, c, val, bg, fg, bold=True, sz=10, ha="left", fmt=None, wrap=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = fill(bg); cell.font = fnt(fg, bold, sz)
        cell.alignment = aln(ha, "center", wrap)
        if fmt: cell.number_format = fmt
        return cell

    # Fila 1 titulo
    ws.merge_cells("A1:J1")
    sc(1,1,f"AUDITORIA  |  {sheet_name}  |  {fecha}","1C1B2E","F2F2F2",True,12)
    ws.row_dimensions[1].height=22

    # Fila 2 KPIs financieros
    ws.row_dimensions[2].height=18
    kpis2=[(1,"TOTAL CALCULADO",f"${tot:,.0f}"),(3,"AREA",f"{area:,.0f} m2"),
           (5,"COSTO / m2",f"${tot/area:,.0f}"),(7,"TOTAL + AIU 15%",f"${tot*1.15:,.0f}"),
           (9,"TOTAL/m2 + AIU",f"${tot*1.15/area:,.0f}")]
    for col,lbl,val in kpis2:
        sc(2,col,lbl,"595959","F2F2F2",True,9)
        sc(2,col+1,val,"D9D9D9","006939",True,10,"right")

    # Fila 3 errores
    ws.row_dimensions[3].height=18
    err_kpis=[(1,"ERRORES TOTALES",resumen["n_err"],"595959"),
              (3,"CRITICOS",resumen["criticos"],"C00000" if resumen["criticos"] else "375623"),
              (5,"FORMULA",resumen["formulas"],"E26B0A" if resumen["formulas"] else "375623"),
              (7,"EXTRA COL",resumen["extras"],"B8860B" if resumen["extras"] else "375623"),
              (9,"PUNTAJE",f"{resumen['puntaje']}/100","375623" if resumen["puntaje"]>=80 else "E26B0A" if resumen["puntaje"]>=60 else "C00000")]
    for col,lbl,val,bg in err_kpis:
        sc(3,col,lbl,"595959","F2F2F2",True,9)
        sc(3,col+1,str(val),"F2F2F2" if bg!="D9D9D9" else "252525",bg,True,11,"right")

    # Fila 4 leyenda semaforo
    ws.row_dimensions[4].height=16
    ley=[(1,"SEMAFORO:","F2F2F2","595959"),(2,"ROJO = Valor absurdo","F2F2F2","C00000"),
         (4,"NARANJA = Error formula","F2F2F2","E26B0A"),(6,"AMARILLO = Col extra","252525","FFD700"),
         (8,"VERDE = Valor correcto","F2F2F2","375623")]
    for col,txt,fg_c,bg_c in ley:
        sc(4,col,txt,bg_c,fg_c,True,9)

    # Fila 5 separador
    ws.row_dimensions[5].height=5
    for c in range(1,11):
        ws.cell(row=5,column=c).fill = fill("1C1B2E")

    # ── PRESUPUESTO CON ESTILOS ORIGINALES ────────────────────
    for i, f in enumerate(all_rows):
        xlsx_row = i + 1 + OFFSET
        vals = (f or [])[:8]
        while len(vals) < 8: vals.append(None)
        errs = errMap.get(i, [])

        # Columnas de auditoria (I y J)
        if i == 0:
            vals.append("AUDITORIA - ERROR DETECTADO"); vals.append("VALOR CORRECTO")
        elif errs:
            msgs = " | ".join(e["msg"] for e in errs)
            corr = next((e["correcto"] for e in errs if e.get("correcto") is not None), None)
            vals.append(msgs); vals.append(corr if corr is not None else "Revisar")
        else:
            vals.append(None); vals.append(None)

        ws.append(vals)
        ws.row_dimensions[xlsx_row].height = 15

        # Aplica estilos originales celda por celda
        for c in range(1, 9):
            cell = ws.cell(row=xlsx_row, column=c)
            sty = estilos_orig.get((i+1, c))  # estilos_orig usa 1-based
            if sty:
                if sty["bg"]:
                    cell.fill = fill(sty["bg"])
                if sty["fg"]:
                    cell.font = Font(color="FF"+sty["fg"].upper(), bold=sty.get("bold",False),
                                     italic=sty.get("italic",False), size=sty.get("sz",10), name="Calibri")
                cell.alignment = aln(sty.get("h","left") or "left", sty.get("v","center") or "center", sty.get("wrap",False) or False)
                if sty.get("fmt") and sty["fmt"] != "General":
                    cell.number_format = sty["fmt"]

        # Semaforo: colorea SOLO las celdas con error (encima del estilo original)
        for e in errs:
            if e["col"] < 8:
                ec = ws.cell(row=xlsx_row, column=e["col"]+1)
                if e["tipo"] == "absurdo": ec.fill=F_ROJO;    ec.font=FNT_ROJO
                elif e["tipo"]=="formula":  ec.fill=F_NARANJA; ec.font=FNT_NARANJA
                else:                       ec.fill=F_AMARILLO;ec.font=FNT_AMARILLO

        # Columna I — descripcion del error
        if errs:
            ac = ws.cell(row=xlsx_row, column=9)
            mt = "absurdo" if any(e["tipo"]=="absurdo" for e in errs) else \
                 "formula" if any(e["tipo"]=="formula" for e in errs) else "extra"
            if mt=="absurdo": ac.fill=F_ROJO;    ac.font=fnt("FFFFFF",True,9)
            elif mt=="formula":ac.fill=F_NARANJA; ac.font=fnt("FFFFFF",True,9)
            else:              ac.fill=F_AMARILLO;ac.font=fnt("252525",True,9)
            ac.alignment=aln("left","center",True)

            # Columna J — valor correcto en verde
            corr = next((e.get("correcto") for e in errs if e.get("correcto") is not None), None)
            if corr is not None:
                jc = ws.cell(row=xlsx_row, column=10)
                jc.value = corr; jc.fill=F_VERDE_OK; jc.font=FNT_VERDE_OK
                jc.alignment=aln("right","center"); jc.number_format="#,##0"

    # Anchos (respeta proporciones originales)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 7
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["J"].width = 18

    ws.freeze_panes = "A7"  # Congela resumen + encabezado presupuesto

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── ENDPOINT PRINCIPAL ────────────────────────────────────────
@app.post("/api/auditar")
async def auditar(
    file: UploadFile = File(...),
    hoja: str = Form(""),
    region: str = Form("bogota"),
    tipo: str = Form("retail_mediano"),
):
    ext = os.path.splitext(file.filename or "")[1].lower()
    content = await file.read()

    # Guarda temporalmente
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(content); tmp_path = tmp.name

    try:
        if ext == ".numbers":
            # Lee con numbers_parser
            from numbers_parser import Document
            doc = Document(tmp_path)
            sheet_names = [s.name for s in doc.sheets]
            if not hoja or hoja not in sheet_names:
                hoja = next((s for s in sheet_names if "obra" in s.lower() and "demol" not in s.lower()), sheet_names[0])

            all_rows = []
            for sheet in doc.sheets:
                if sheet.name != hoja: continue
                for row in sheet.tables[0].iter_rows():
                    rv = []
                    for cell in row:
                        try: rv.append(cell.value)
                        except: rv.append(None)
                    all_rows.append(rv)
                break

            estilos_orig = leer_estilos_numbers(tmp_path, hoja)

        elif ext in (".xlsx", ".xls"):
            wb_orig = openpyxl.load_workbook(tmp_path, data_only=True)
            sheet_names = wb_orig.sheetnames
            if not hoja or hoja not in sheet_names:
                hoja = next((s for s in sheet_names if "obra" in s.lower() and "demol" not in s.lower()), sheet_names[0])

            ws_orig = wb_orig[hoja]
            all_rows = [[cell.value for cell in row] for row in ws_orig.iter_rows()]
            estilos_orig = leer_estilos_xlsx(ws_orig)
        else:
            raise HTTPException(400, "Formato no soportado. Usa .xlsx o .numbers")

        errMap  = detectar_errores(all_rows)
        resumen = calcular_resumen(all_rows, errMap)
        buf     = escribir_excel(all_rows, estilos_orig, errMap, resumen, hoja)

        nombre  = f"Auditoria_{hoja.replace(' ','_')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nombre}"',
                     "X-Errores": str(resumen["n_err"]),
                     "X-Puntaje": str(resumen["puntaje"])}
        )
    finally:
        os.unlink(tmp_path)

@app.get("/api/hojas")
async def listar_hojas(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    content = await file.read()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(content); tmp_path = tmp.name
    try:
        if ext == ".numbers":
            from numbers_parser import Document
            doc = Document(tmp_path)
            return {"hojas": [s.name for s in doc.sheets]}
        else:
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            return {"hojas": wb.sheetnames}
    finally:
        os.unlink(tmp_path)

@app.get("/")
def root(): return {"status": "ok", "msg": "Auditoria API lista"}

# ── Sirve el frontend React compilado ────────────────────────
import os
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

static_dir = os.path.join(os.path.dirname(__file__), "static")

@app.get("/")
async def root():
    index = os.path.join(static_dir, "index.html")
    if os.path.exists(index):
        return FileResponse(index)
    return {"status": "API ok - frontend no compilado"}

@app.get("/{full_path:path}")
async def spa(full_path: str):
    # Sirve assets estaticos si existen
    asset = os.path.join(static_dir, full_path)
    if os.path.exists(asset) and os.path.isfile(asset):
        return FileResponse(asset)
    # Todo lo demas -> index.html (SPA routing)
    index = os.path.join(static_dir, "index.html")
    if os.path.exists(index):
        return FileResponse(index)
    return {"msg": "no encontrado"}
